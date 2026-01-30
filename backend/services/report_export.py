"""
Report Export Service.

Generates PDF and Excel reports for backtesting results:
- Comprehensive metrics summary
- Equity curve charts
- Trade list tables
- Performance analysis
"""

import io
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Optional

logger = logging.getLogger(__name__)


@dataclass
class BacktestReportData:
    """Data container for backtest report."""

    backtest_id: str
    symbol: str
    interval: str
    start_date: str
    end_date: str
    strategy_type: str
    strategy_params: dict[str, Any]
    initial_capital: float
    final_capital: float
    total_return_pct: float
    total_trades: int
    winning_trades: int
    losing_trades: int
    win_rate: float
    profit_factor: float
    sharpe_ratio: float
    sortino_ratio: float
    max_drawdown: float
    avg_trade_pnl: float
    avg_win: float
    avg_loss: float
    largest_win: float
    largest_loss: float
    equity_curve: list[dict[str, Any]]
    trades: list[dict[str, Any]]
    monthly_returns: Optional[list[dict[str, Any]]] = None
    additional_metrics: Optional[dict[str, Any]] = None


class ReportExportService:
    """
    Service for exporting backtest reports.

    Supports PDF and Excel formats with comprehensive metrics.
    """

    def __init__(self):
        """Initialize report export service."""
        self._check_dependencies()

    def _check_dependencies(self) -> dict[str, bool]:
        """Check which export libraries are available."""
        deps = {
            "reportlab": False,
            "openpyxl": False,
            "xlsxwriter": False,
            "matplotlib": False,
        }

        try:
            import reportlab  # noqa: F401

            deps["reportlab"] = True
        except ImportError:
            pass

        try:
            import openpyxl  # noqa: F401

            deps["openpyxl"] = True
        except ImportError:
            pass

        try:
            import xlsxwriter  # noqa: F401

            deps["xlsxwriter"] = True
        except ImportError:
            pass

        try:
            import matplotlib  # noqa: F401

            deps["matplotlib"] = True
        except ImportError:
            pass

        return deps

    def generate_excel_report(self, data: BacktestReportData) -> bytes:
        """
        Generate Excel report.

        Args:
            data: Backtest report data

        Returns:
            Excel file as bytes
        """
        try:
            import xlsxwriter  # noqa: F401
        except ImportError:
            try:
                import openpyxl  # noqa: F401

                return self._generate_excel_openpyxl(data)
            except ImportError:
                raise RuntimeError(
                    "No Excel library available. Install xlsxwriter or openpyxl."
                )

        return self._generate_excel_xlsxwriter(data)

    def _generate_excel_xlsxwriter(self, data: BacktestReportData) -> bytes:
        """Generate Excel using xlsxwriter."""
        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})

        # Formats
        header_format = workbook.add_format(
            {
                "bold": True,
                "bg_color": "#1a1a2e",
                "font_color": "white",
                "border": 1,
            }
        )
        number_format = workbook.add_format({"num_format": "#,##0.00"})
        pct_format = workbook.add_format({"num_format": "0.00%"})
        currency_format = workbook.add_format({"num_format": "$#,##0.00"})
        positive_format = workbook.add_format(
            {"num_format": "$#,##0.00", "font_color": "green"}
        )
        negative_format = workbook.add_format(
            {"num_format": "$#,##0.00", "font_color": "red"}
        )
        date_format = workbook.add_format({"num_format": "yyyy-mm-dd hh:mm"})  # noqa: F841

        # --- Summary Sheet ---
        summary = workbook.add_worksheet("Summary")
        summary.set_column("A:A", 25)
        summary.set_column("B:B", 20)

        summary.write("A1", "Backtest Report", header_format)
        summary.write(
            "B1", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"), header_format
        )

        row = 3
        summary_data = [
            ("Backtest ID", data.backtest_id),
            ("Symbol", data.symbol),
            ("Interval", data.interval),
            ("Period", f"{data.start_date} to {data.end_date}"),
            ("Strategy", data.strategy_type),
            ("", ""),
            ("CAPITAL", ""),
            ("Initial Capital", data.initial_capital),
            ("Final Capital", data.final_capital),
            ("Total Return", f"{data.total_return_pct:.2f}%"),
            ("", ""),
            ("TRADES", ""),
            ("Total Trades", data.total_trades),
            ("Winning Trades", data.winning_trades),
            ("Losing Trades", data.losing_trades),
            ("Win Rate", f"{data.win_rate:.2f}%"),
            ("", ""),
            ("PERFORMANCE", ""),
            ("Profit Factor", data.profit_factor),
            ("Sharpe Ratio", data.sharpe_ratio),
            ("Sortino Ratio", data.sortino_ratio),
            ("Max Drawdown", f"{data.max_drawdown:.2f}%"),
            ("", ""),
            ("TRADE METRICS", ""),
            ("Avg Trade P&L", data.avg_trade_pnl),
            ("Avg Win", data.avg_win),
            ("Avg Loss", data.avg_loss),
            ("Largest Win", data.largest_win),
            ("Largest Loss", data.largest_loss),
        ]

        for label, value in summary_data:
            summary.write(row, 0, label)
            if isinstance(value, float) and "Capital" in label:
                summary.write(row, 1, value, currency_format)
            elif isinstance(value, float):
                summary.write(row, 1, value, number_format)
            else:
                summary.write(row, 1, value)
            row += 1

        # Strategy parameters
        row += 2
        summary.write(row, 0, "Strategy Parameters", header_format)
        summary.write(row, 1, "", header_format)
        row += 1
        for key, value in data.strategy_params.items():
            summary.write(row, 0, key)
            summary.write(row, 1, str(value))
            row += 1

        # --- Trades Sheet ---
        trades_sheet = workbook.add_worksheet("Trades")
        trade_headers = [
            "Trade #",
            "Type",
            "Entry Time",
            "Exit Time",
            "Entry Price",
            "Exit Price",
            "Qty",
            "P&L",
            "P&L %",
            "Duration",
        ]

        for col, header in enumerate(trade_headers):
            trades_sheet.write(0, col, header, header_format)

        trades_sheet.set_column("A:A", 8)
        trades_sheet.set_column("B:B", 8)
        trades_sheet.set_column("C:D", 18)
        trades_sheet.set_column("E:F", 12)
        trades_sheet.set_column("G:I", 12)
        trades_sheet.set_column("J:J", 15)

        for row, trade in enumerate(data.trades, start=1):
            trades_sheet.write(row, 0, row)
            trades_sheet.write(row, 1, trade.get("side", trade.get("type", "N/A")))
            trades_sheet.write(row, 2, trade.get("entry_time", ""))
            trades_sheet.write(row, 3, trade.get("exit_time", ""))
            trades_sheet.write(row, 4, trade.get("entry_price", 0), number_format)
            trades_sheet.write(row, 5, trade.get("exit_price", 0), number_format)
            trades_sheet.write(
                row, 6, trade.get("qty", trade.get("size", 0)), number_format
            )

            pnl = trade.get("pnl", trade.get("profit", 0))
            fmt = positive_format if pnl >= 0 else negative_format
            trades_sheet.write(row, 7, pnl, fmt)

            pnl_pct = trade.get("pnl_pct", trade.get("return_pct", 0))
            trades_sheet.write(row, 8, f"{pnl_pct:.2f}%")
            trades_sheet.write(row, 9, trade.get("duration", "N/A"))

        # --- Equity Curve Sheet ---
        if data.equity_curve:
            equity_sheet = workbook.add_worksheet("Equity Curve")
            equity_headers = ["Timestamp", "Equity", "Drawdown %"]

            for col, header in enumerate(equity_headers):
                equity_sheet.write(0, col, header, header_format)

            equity_sheet.set_column("A:A", 20)
            equity_sheet.set_column("B:C", 15)

            for row, point in enumerate(data.equity_curve, start=1):
                equity_sheet.write(
                    row, 0, point.get("timestamp", point.get("time", ""))
                )
                equity_sheet.write(
                    row, 1, point.get("equity", point.get("value", 0)), currency_format
                )
                equity_sheet.write(row, 2, point.get("drawdown", 0), pct_format)

            # Add equity chart
            chart = workbook.add_chart({"type": "line"})
            chart.add_series(
                {
                    "name": "Equity",
                    "categories": f"='Equity Curve'!$A$2:$A${len(data.equity_curve) + 1}",
                    "values": f"='Equity Curve'!$B$2:$B${len(data.equity_curve) + 1}",
                    "line": {"color": "#00D4AA", "width": 2},
                }
            )
            chart.set_title({"name": "Equity Curve"})
            chart.set_x_axis({"name": "Time"})
            chart.set_y_axis({"name": "Equity ($)"})
            chart.set_size({"width": 720, "height": 400})
            equity_sheet.insert_chart("E2", chart)

        # --- Monthly Returns Sheet ---
        if data.monthly_returns:
            monthly_sheet = workbook.add_worksheet("Monthly Returns")
            monthly_headers = ["Month", "Return %", "Trades", "Win Rate %"]

            for col, header in enumerate(monthly_headers):
                monthly_sheet.write(0, col, header, header_format)

            for row, month in enumerate(data.monthly_returns, start=1):
                monthly_sheet.write(row, 0, month.get("month", ""))
                ret = month.get("return", 0)
                fmt = positive_format if ret >= 0 else negative_format
                monthly_sheet.write(row, 1, f"{ret:.2f}%")
                monthly_sheet.write(row, 2, month.get("trades", 0))
                monthly_sheet.write(row, 3, f"{month.get('win_rate', 0):.2f}%")

        workbook.close()
        output.seek(0)
        return output.read()

    def _generate_excel_openpyxl(self, data: BacktestReportData) -> bytes:
        """Generate Excel using openpyxl (fallback)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Header style
        header_fill = PatternFill(
            start_color="1a1a2e", end_color="1a1a2e", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # Summary data
        ws["A1"] = "Backtest Report"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill

        summary_rows = [
            ("Backtest ID", data.backtest_id),
            ("Symbol", data.symbol),
            ("Interval", data.interval),
            ("Strategy", data.strategy_type),
            ("Initial Capital", f"${data.initial_capital:,.2f}"),
            ("Final Capital", f"${data.final_capital:,.2f}"),
            ("Total Return", f"{data.total_return_pct:.2f}%"),
            ("Total Trades", data.total_trades),
            ("Win Rate", f"{data.win_rate:.2f}%"),
            ("Profit Factor", f"{data.profit_factor:.2f}"),
            ("Sharpe Ratio", f"{data.sharpe_ratio:.2f}"),
            ("Max Drawdown", f"{data.max_drawdown:.2f}%"),
        ]

        for row, (label, value) in enumerate(summary_rows, start=3):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)

        # Trades sheet
        trades_ws = wb.create_sheet("Trades")
        headers = [
            "#",
            "Type",
            "Entry Time",
            "Exit Time",
            "Entry Price",
            "Exit Price",
            "P&L",
        ]

        for col, header in enumerate(headers, start=1):
            cell = trades_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, trade in enumerate(data.trades, start=2):
            trades_ws.cell(row=row, column=1, value=row - 1)
            trades_ws.cell(row=row, column=2, value=trade.get("side", ""))
            trades_ws.cell(row=row, column=3, value=trade.get("entry_time", ""))
            trades_ws.cell(row=row, column=4, value=trade.get("exit_time", ""))
            trades_ws.cell(row=row, column=5, value=trade.get("entry_price", 0))
            trades_ws.cell(row=row, column=6, value=trade.get("exit_price", 0))
            trades_ws.cell(row=row, column=7, value=trade.get("pnl", 0))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_pdf_report(self, data: BacktestReportData) -> bytes:
        """
        Generate PDF report.

        Args:
            data: Backtest report data

        Returns:
            PDF file as bytes
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            raise RuntimeError("reportlab not installed. Run: pip install reportlab")

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=20,
            textColor=colors.HexColor("#1a1a2e"),
            spaceAfter=20,
        )
        heading_style = ParagraphStyle(
            "Heading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#00D4AA"),
            spaceBefore=15,
            spaceAfter=10,
        )
        normal_style = styles["Normal"]

        elements = []

        # Title
        elements.append(Paragraph("Backtest Report", title_style))
        elements.append(
            Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
                normal_style,
            )
        )
        elements.append(Spacer(1, 20))

        # Overview section
        elements.append(Paragraph("Overview", heading_style))
        overview_data = [
            ["Symbol", data.symbol, "Strategy", data.strategy_type],
            [
                "Interval",
                data.interval,
                "Period",
                f"{data.start_date} - {data.end_date}",
            ],
            [
                "Initial Capital",
                f"${data.initial_capital:,.2f}",
                "Final Capital",
                f"${data.final_capital:,.2f}",
            ],
        ]
        overview_table = Table(
            overview_data, colWidths=[1.5 * inch, 2 * inch, 1.5 * inch, 2 * inch]
        )
        overview_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f0f0")),
                    ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f0f0f0")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                    ("PADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(overview_table)
        elements.append(Spacer(1, 20))

        # Performance metrics
        elements.append(Paragraph("Performance Metrics", heading_style))

        # Color returns
        return_color = colors.green if data.total_return_pct >= 0 else colors.red

        metrics_data = [
            [
                "Total Return",
                f"{data.total_return_pct:+.2f}%",
                "Win Rate",
                f"{data.win_rate:.2f}%",
            ],
            [
                "Total Trades",
                str(data.total_trades),
                "Profit Factor",
                f"{data.profit_factor:.2f}",
            ],
            [
                "Winning Trades",
                str(data.winning_trades),
                "Sharpe Ratio",
                f"{data.sharpe_ratio:.2f}",
            ],
            [
                "Losing Trades",
                str(data.losing_trades),
                "Sortino Ratio",
                f"{data.sortino_ratio:.2f}",
            ],
            [
                "Max Drawdown",
                f"{data.max_drawdown:.2f}%",
                "Avg Trade P&L",
                f"${data.avg_trade_pnl:,.2f}",
            ],
        ]
        metrics_table = Table(
            metrics_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch]
        )
        metrics_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f0f0")),
                    ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f0f0f0")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                    ("TEXTCOLOR", (1, 0), (1, 0), return_color),
                    ("PADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(metrics_table)
        elements.append(Spacer(1, 20))

        # Strategy parameters
        elements.append(Paragraph("Strategy Parameters", heading_style))
        params_data = [[k, str(v)] for k, v in data.strategy_params.items()]
        if params_data:
            params_table = Table(params_data, colWidths=[2 * inch, 3 * inch])
            params_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f0f0")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("PADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(params_table)
        elements.append(Spacer(1, 20))

        # Trade summary
        elements.append(Paragraph("Trade Summary", heading_style))
        trade_summary = [
            ["Metric", "Value"],
            ["Average Win", f"${data.avg_win:,.2f}"],
            ["Average Loss", f"${data.avg_loss:,.2f}"],
            ["Largest Win", f"${data.largest_win:,.2f}"],
            ["Largest Loss", f"${data.largest_loss:,.2f}"],
        ]
        trade_table = Table(trade_summary, colWidths=[2 * inch, 2 * inch])
        trade_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f0f0f0")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("PADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(trade_table)

        # Page break before trades list
        elements.append(PageBreak())
        elements.append(Paragraph("Trade List", heading_style))

        # Trades table (first 50)
        trades_header = ["#", "Type", "Entry", "Exit", "Entry $", "Exit $", "P&L"]
        trades_data = [trades_header]

        for i, trade in enumerate(data.trades[:50], start=1):
            pnl = trade.get("pnl", trade.get("profit", 0))
            trades_data.append(
                [
                    str(i),
                    trade.get("side", trade.get("type", "")),
                    trade.get("entry_time", "")[:16] if trade.get("entry_time") else "",
                    trade.get("exit_time", "")[:16] if trade.get("exit_time") else "",
                    f"${trade.get('entry_price', 0):,.2f}",
                    f"${trade.get('exit_price', 0):,.2f}",
                    f"${pnl:+,.2f}",
                ]
            )

        trades_table = Table(
            trades_data,
            colWidths=[
                0.4 * inch,
                0.6 * inch,
                1.2 * inch,
                1.2 * inch,
                1 * inch,
                1 * inch,
                1 * inch,
            ],
        )

        # Color P&L column
        table_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]

        # Color P&L cells
        for i, trade in enumerate(data.trades[:50], start=1):
            pnl = trade.get("pnl", trade.get("profit", 0))
            color = colors.green if pnl >= 0 else colors.red
            table_style.append(("TEXTCOLOR", (6, i), (6, i), color))

        trades_table.setStyle(TableStyle(table_style))
        elements.append(trades_table)

        if len(data.trades) > 50:
            elements.append(
                Paragraph(f"... and {len(data.trades) - 50} more trades", normal_style)
            )

        doc.build(elements)
        output.seek(0)
        return output.read()

    def generate_csv_report(self, data: BacktestReportData) -> bytes:
        """
        Generate CSV report (trades only).

        Args:
            data: Backtest report data

        Returns:
            CSV file as bytes
        """
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(
            [
                "Trade #",
                "Type",
                "Entry Time",
                "Exit Time",
                "Entry Price",
                "Exit Price",
                "Quantity",
                "P&L",
                "P&L %",
            ]
        )

        # Trades
        for i, trade in enumerate(data.trades, start=1):
            writer.writerow(
                [
                    i,
                    trade.get("side", trade.get("type", "")),
                    trade.get("entry_time", ""),
                    trade.get("exit_time", ""),
                    trade.get("entry_price", 0),
                    trade.get("exit_price", 0),
                    trade.get("qty", trade.get("size", 0)),
                    trade.get("pnl", trade.get("profit", 0)),
                    trade.get("pnl_pct", trade.get("return_pct", 0)),
                ]
            )

        return output.getvalue().encode("utf-8")


# Global instance
_report_service: Optional[ReportExportService] = None


def get_report_service() -> ReportExportService:
    """Get or create report export service."""
    global _report_service
    if _report_service is None:
        _report_service = ReportExportService()
    return _report_service
